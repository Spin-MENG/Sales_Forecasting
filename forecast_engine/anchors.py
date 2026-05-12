import openpyxl

from .validation import BusinessValidationError


def month_index(month_text):
    year, month = str(month_text).split("-")
    return int(year) * 12 + int(month)


def extract_country_totals(fn, label, excel_config, start_month=None, end_month=None):
    wb = openpyxl.load_workbook(fn, data_only=True)
    sheet_name = excel_config.get("sheet_name", "销量")
    if sheet_name not in wb.sheetnames:
        raise BusinessValidationError(f"{label} 文件中没有「{sheet_name}」工作表，无法读取销量。")
    ws = wb[sheet_name]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    country_header = excel_config.get("country_column", "国家")
    store_header = excel_config.get("store_column", "店铺")
    missing_headers = [h for h in (country_header, store_header) if h not in headers]
    if missing_headers:
        raise BusinessValidationError(
            f"{label} 文件缺少字段「{' / '.join(missing_headers)}」，请确认表头包含「{country_header}」和「{store_header}」。"
        )
    country_col = headers.index(country_header) + 1
    store_col = headers.index(store_header) + 1
    month_cols = []
    for i, h in enumerate(headers):
        if not (isinstance(h, str) and len(h) == 7 and h[:4].isdigit()):
            continue
        if start_month and h < start_month:
            continue
        if end_month and h > end_month:
            continue
        month_cols.append((i + 1, h))
    if not month_cols:
        raise BusinessValidationError(f"{label} 在配置窗口 {start_month} ~ {end_month} 内没有可用月份列。")
    for (_, prev_month), (_, next_month) in zip(month_cols, month_cols[1:]):
        if month_index(next_month) - month_index(prev_month) != 1:
            raise BusinessValidationError(
                f"{label} 的月份字段不连续：{prev_month} 后面接到 {next_month}，请补齐缺失月份或调整 start_month / end_month。"
            )

    country_totals = {}
    for r in range(2, ws.max_row + 1):
        country = ws.cell(row=r, column=country_col).value
        store = ws.cell(row=r, column=store_col).value
        total = sum((ws.cell(row=r, column=c).value or 0) for c, _ in month_cols)
        if total > 0:
            country_totals[country] = country_totals.get(country, 0) + total
    us = country_totals.get(excel_config.get("us_country", "美国"), 0)
    de = country_totals.get(excel_config.get("de_country", "德国"), 0)
    grand = sum(country_totals.values())
    if grand <= 0:
        raise BusinessValidationError(f"{label} 在配置窗口内销量全为 0，无法计算 DE/非美占比。")
    if de <= 0:
        raise BusinessValidationError(f"{label} 文件中没有德国销量，无法计算 DE/非美占比。")
    non_us = grand - us
    if non_us <= 0:
        raise BusinessValidationError(f"{label} 非美销量为 0，无法计算 DE/非美占比。")
    de_pct_global = de / grand if grand > 0 else 0
    de_pct_non_us = de / non_us if non_us > 0 else 0
    min_share = float(excel_config.get("de_non_us_share_min", 0.03))
    max_share = float(excel_config.get("de_non_us_share_max", 0.60))
    if de_pct_non_us < min_share or de_pct_non_us > max_share:
        raise BusinessValidationError(
            f"{label} 的 DE/非美占比为 {de_pct_non_us * 100:.1f}%，超出合理范围 "
            f"{min_share * 100:.0f}%~{max_share * 100:.0f}%；请检查德国、美国或其他国家销量是否漏填。"
        )
    print(f"\n{label}:")
    print(f"  月份数: {len(month_cols)} ({month_cols[0][1]} ~ {month_cols[-1][1]})")
    us_text = f"{us:,.0f} ({us / grand * 100:.1f}%)" if us > 0 else "未提供/为 0（按无美国数据继续计算）"
    print(f"  全球总销量: {grand:,.0f}, 美国: {us_text}, 非美: {non_us:,.0f}")
    print(f"  德国销量: {de:,.0f}")
    print(f"  DE / 全球   = {de_pct_global * 100:.1f}%")
    print(f"  DE / 非美   = {de_pct_non_us * 100:.1f}%")
    return {
        "label": label,
        "us": us,
        "non_us": non_us,
        "de": de,
        "grand": grand,
        "de_pct_global": de_pct_global,
        "de_pct_non_us": de_pct_non_us,
        "n_months": len(month_cols),
        "country_totals": country_totals,
        "start_month": month_cols[0][1],
        "end_month": month_cols[-1][1],
    }
